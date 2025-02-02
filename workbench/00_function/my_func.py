from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side,Alignment
def write_excel(excel_path, sheet_name='sheet_py'):
  # エクセルファイルが開いていたら処理を終了
  try:
    wb = load_workbook(excel_path)
    ws = wb.active
    wb.close()
  except:
    print('エクセルファイルが開いているため処理を終了します')
    pass
  # エクセルファイルの読み込み
  wb = load_workbook(excel_path)
  # シートが存在しない場合は新規作成
  if sheet_name not in wb.sheetnames:
    wb.create_sheet(sheet_name)
  # 対象のシートを取得
  ws = wb[sheet_name]
  # 罫線なしのスタイルを作成
  no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
  # 左寄せ & 上詰め の配置を定義
  alignment = Alignment(horizontal="left", vertical="top")
  for row in ws.iter_rows():
    for cell in row:
      cell.font = Font(name='メイリオ', size=9)
      cell.border = no_border
      cell.alignment = alignment
  # エクセルファイルの保存
  wb.save(excel_path)
def write_dataframe_to_excel(df, excel_path, sheet_name='sheet_py', start_row=1, start_col=1):
  """
  pandasのデータフレームを指定されたExcelファイルの特定のシートとセルに書き込む関数
  :param df: pandasのデータフレーム
  :param excel_path: 出力先Excelファイルのパス
  :param sheet_name: 書き込むシートの名前(デフォルト：'sheet_py')
  :param start_row: データを書き込む開始行(デフォルト：1)
  :param start_col: データを書き込む開始列(デフォルト：1)
  """
  # エクセルファイルが開いていたら処理を終了
  try:
    wb = load_workbook(excel_path)
    ws = wb.active
    wb.close()
  except:
    print('エクセルファイルが開いているため処理を終了します')
    pass
  # Excelファイルの読み込み
  try:
    wb = load_workbook(excel_path)
  except FileNotFoundError:
    print("ファイルが見つかりません")
    pass
  # シートが存在しない場合は新規作成
  if sheet_name not in wb.sheetnames:
    wb.create_sheet(sheet_name)
  # ワークシートの取得
  ws = wb[sheet_name]
  # ヘッダーを設定 (太字)
  for col_num, column_name in enumerate(df.columns, start=start_col):
    ws.cell(row=start_row, column=col_num, value=column_name)
  # データフレームの各行をExcelに書き込む
  for i, row in enumerate(df.itertuples(), start=start_row + 1):
    for j, value in enumerate(row[1:], start=start_col):  # row[1:]でインデックスをスキップ
      ws.cell(row=i, column=j, value=value)
  # Excelファイルの保存
  wb.save(excel_path)
  print(f"データが {excel_path} に保存されました。")